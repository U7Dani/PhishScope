import os
import email
import hashlib
import pandas as pd
import tldextract
import dns.resolver
import requests
import json
import extract_msg
from bs4 import BeautifulSoup
from email import policy
from email.parser import BytesParser
from sentence_transformers import SentenceTransformer, util
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from config_loader import cargar_config

# Cargar configuración desde archivo externo
config = cargar_config()
EMAIL_DIR = config["EMAIL_DIR"]
SALIDA_XLSX = config["SALIDA_XLSX"]
ACORTADORES = config["ACORTADORES"]
EXT_PELIGROSAS = config["EXT_PELIGROSAS"]
PATRONES = config["PATRONES"]

# Carga modelo de embeddings
modelo_nlp = SentenceTransformer('sentence-transformers/paraphrase-MiniLM-L6-v2')

def calcular_similitud(texto):
    resultados = {k: 0 for k in PATRONES}
    emb_texto = modelo_nlp.encode(texto, convert_to_tensor=True)
    for categoria, frases in PATRONES.items():
        for frase in frases:
            emb_patron = modelo_nlp.encode(frase, convert_to_tensor=True)
            sim = util.cos_sim(emb_texto, emb_patron).item()
            if sim > 0.7:
                resultados[categoria] += 1
    return resultados

def extraer_urls(html):
    soup = BeautifulSoup(html, "html.parser")
    return [a['href'] for a in soup.find_all('a', href=True)]

def resolver_redireccion(url):
    try:
        return requests.head(url, allow_redirects=True, timeout=5).url
    except:
        return url

def verificar_auth(domain):
    resultados = {"spf": "❌", "dkim": "❌", "dmarc": "❌"}
    try:
        if any("v=spf1" in r.to_text() for r in dns.resolver.resolve(domain, "TXT")):
            resultados["spf"] = "✔️"
    except: pass
    try:
        if dns.resolver.resolve(f"_dmarc.{domain}", "TXT"):
            resultados["dmarc"] = "✔️"
    except: pass
    try:
        if dns.resolver.resolve(f"selector1._domainkey.{domain}", "TXT"):
            resultados["dkim"] = "✔️"
    except: pass
    return resultados

def analizar_adjuntos(msg):
    adjuntos = []
    for part in msg.walk():
        if part.get_content_disposition() == "attachment":
            filename = part.get_filename()
            ext = os.path.splitext(filename)[-1].lower()
            if ext in EXT_PELIGROSAS:
                contenido = part.get_payload(decode=True)
                sha256 = hashlib.sha256(contenido).hexdigest()
                adjuntos.append((filename, sha256))
    return adjuntos

def procesar_msg(filepath):
    msg = extract_msg.Message(filepath)
    remitente = msg.sender or ""
    reply_to = msg.reply_to or ""
    subject = msg.subject or ""
    cuerpo = msg.body or ""
    adjuntos = [att.longFilename for att in msg.attachments]
    return remitente, reply_to, subject, cuerpo, adjuntos

def calcular_phishscore(sim_nlp, urls_info, adjuntos, auth):
    score = 0
    score += sum(sim_nlp.values()) * 10
    score += sum(20 for a in urls_info if a[3] or tldextract.extract(a[1]).top_domain_under_public_suffix in ACORTADORES)
    score += 15 * len(adjuntos)
    score += 15 if auth["reply_diff"] else 0
    score += -10 if auth["spf"] == auth["dkim"] == auth["dmarc"] == "✔️" else 10
    return min(score, 100)

# Procesar correos
resultados = []
for archivo in os.listdir(EMAIL_DIR):
    ruta = os.path.join(EMAIL_DIR, archivo)

    if archivo.endswith(".eml"):
        with open(ruta, "rb") as f:
            msg = BytesParser(policy=policy.default).parse(f)
        remitente = msg.get("From", "")
        reply_to = msg.get("Reply-To", "")
        subject = msg.get("Subject", "")
        cuerpo = msg.get_body(preferencelist=("html", "plain")).get_content()
        adjuntos = analizar_adjuntos(msg)

    elif archivo.endswith(".msg"):
        remitente, reply_to, subject, cuerpo, adjuntos = procesar_msg(ruta)

    else:
        continue

    urls = extraer_urls(cuerpo)
    urls_info = [(u, resolver_redireccion(u), u, tldextract.extract(u).top_domain_under_public_suffix != tldextract.extract(resolver_redireccion(u)).top_domain_under_public_suffix) for u in urls]
    sim_nlp = calcular_similitud(subject + " " + cuerpo)
    domain = tldextract.extract(remitente).top_domain_under_public_suffix
    auth = verificar_auth(domain)
    auth["reply_diff"] = reply_to and reply_to not in remitente
    score = calcular_phishscore(sim_nlp, urls_info, adjuntos, auth)

    if score >= 70:
        riesgo = "🟥 Phishing muy probable"
    elif score >= 50:
        riesgo = "🟧 Posible phishing"
    elif score >= 30:
        riesgo = "🟨 Sospechoso"
    else:
        riesgo = "🟩 Limpio"

    resultados.append({
        "archivo": archivo,
        "remitente_nombre": remitente.split("<")[0].strip(),
        "remitente_email": remitente.split("<")[-1].strip(">").strip(),
        "subject": subject,
        "score": score,
        "riesgo": riesgo,
        "enlaces": ", ".join(u[1] for u in urls_info),
        "adjuntos": ", ".join(f"{a[0]} ({a[1][:6]}...)" if isinstance(a, tuple) else a for a in adjuntos),
        "nlp_urgencia": sim_nlp["urgencia"],
        "nlp_autoridad": sim_nlp["autoridad"],
        "nlp_recompensa": sim_nlp["recompensa"],
        "nlp_castigo": sim_nlp["castigo"],
        "spf": auth["spf"],
        "dkim": auth["dkim"],
        "dmarc": auth["dmarc"]
    })

# Exportar a Excel con formato
df = pd.DataFrame(resultados)
df.to_excel(SALIDA_XLSX, index=False)

# Formato condicional
wb = load_workbook(SALIDA_XLSX)
ws = wb.active
col_riesgo = "F"
col_range = f"{col_riesgo}2:{col_riesgo}{ws.max_row}"
colores = [
    ("\"🟥 Phishing muy probable\"", "FF0000"),
    ("\"🟧 Posible phishing\"", "FFA500"),
    ("\"🟨 Sospechoso\"", "FFFF00"),
    ("\"🟩 Limpio\"", "C6EFCE"),
]
for val, color in colores:
    regla = FormulaRule(formula=[f"${col_riesgo}2={val}"], fill=PatternFill(start_color=color, end_color=color, fill_type="solid"))
    ws.conditional_formatting.add(col_range, regla)

# Congelar fila de encabezado y autoajuste de ancho
ws.freeze_panes = "A2"
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max(10, max_length + 2)

wb.save(SALIDA_XLSX)
wb.close()
